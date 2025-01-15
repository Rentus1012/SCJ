# from openpyxl import load_workbook
# from collections import Counter

# # 퀵 소트 함수
# def quick_sort(array):
#     if len(array) <= 1:
#         return array
#     pivot = array[len(array) // 2]  # 중앙값을 피벗으로 선택
#     left = [x for x in array if x < pivot]
#     middle = [x for x in array if x == pivot]
#     right = [x for x in array if x > pivot]
#     return quick_sort(left) + middle + quick_sort(right)

# def check_duplicate_phone_numbers(file_path, sheet_name):
#     # 엑셀 파일 로드
#     wb = load_workbook(file_path)
#     ws = wb[sheet_name]

#     # E열 데이터 가져오기 (E1부터 끝까지)
#     phone_numbers = [cell.value for cell in ws['E'] if cell.value is not None]

#     # 퀵 소트를 이용해 정렬
#     sorted_numbers = quick_sort(phone_numbers)

#     # 전화번호 빈도수 체크 (Counter 사용)
#     phone_count = Counter(sorted_numbers)

#     # 중복된 전화번호를 저장할 리스트
#     duplicates = [phone for phone, count in phone_count.items() if count > 1]

#     # 중복 전화번호가 있으면 출력
#     if duplicates:
#         print("중복된 전화번호가 발견되었습니다:")
#         for phone in duplicates:
#             print(phone)
#     else:
#         print("중복된 전화번호가 없습니다.")

#     # 프로그램 끝까지 진행
#     print("파일 끝까지 확인되었습니다.")

#     while 1:
#         a = input("완료되었습니다. 종료 하실려면 q 키를 누르고 ENTER 키를 눌러 종료해주세요.")

#         if a == "q" or "Q":
#             break

# # 파일 경로 및 시트 이름
# file_path = "C:/Users/seung/바탕 화면/DevDrive/Backend/1111.xlsx"
# sheet_name = "Sheet1"
# check_duplicate_phone_numbers(file_path, sheet_name)


from openpyxl import load_workbook
from collections import Counter
from tkinter import Tk, filedialog
import requests

# 퀵 소트 함수
def quick_sort(array):
    if len(array) <= 1:
        return array
    pivot = array[len(array) // 2]  # 중앙값을 피벗으로 선택
    left = [x for x in array if x < pivot]
    middle = [x for x in array if x == pivot]
    right = [x for x in array if x > pivot]
    return quick_sort(left) + middle + quick_sort(right)

def check_duplicate_phone_numbers():
    # 파일 선택 창 띄우기
    Tk().withdraw()  # Tkinter 기본 창 숨기기
    file_path = filedialog.askopenfilename(
        title="엑셀 파일 선택",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if not file_path:
        print("파일이 선택되지 않았습니다. 프로그램을 종료합니다.")
        return

    # 시트 이름 입력받기
    sheet_name = input("확인할 시트 이름을 입력하세요: ").strip()
    if not sheet_name:
        print("시트 이름이 입력되지 않았습니다. 프로그램을 종료합니다.")
        return

    # 컬럼 이름 입력받기
    column_name = input("확인할 컬럼 이름을 입력하세요 (예: A, B, C, ...): ").strip().upper()
    if not column_name.isalpha() or len(column_name) != 1:
        print("올바른 컬럼 이름을 입력하세요 (예: A, B, C, ...). 프로그램을 종료합니다.")
        return

    try:
        # 엑셀 파일 로드
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' 시트를 찾을 수 없습니다. 프로그램을 종료합니다.")
            return

        ws = wb[sheet_name]

        # 선택한 컬럼 데이터 가져오기
        column_data = [cell.value for cell in ws[column_name] if cell.value is not None]

        # 퀵 소트를 이용해 정렬
        sorted_data = quick_sort(column_data)

        # 데이터 빈도수 체크 (Counter 사용)
        data_count = Counter(sorted_data)

        # 중복된 데이터를 저장할 리스트
        duplicates = [data for data, count in data_count.items() if count > 1]

        # 중복 데이터가 있으면 출력
        if duplicates:
            print("중복된 데이터가 발견되었습니다:")
            for data in duplicates:
                print(data)
        else:
            print("중복된 데이터가 없습니다.")

        # 프로그램 끝까지 진행
        print("파일 끝까지 확인되었습니다.")
    except Exception as e:
        print(f"오류가 발생했습니다: {e}")
        return

    # 종료 대기
    while True:
        a = input("완료되었습니다. 종료하시려면 'q'를 누르고 ENTER 키를 눌러주세요: ").strip().lower()
        if a == 'q':
            break


# 구글 스프레드시트 링크에서 XLSX 파일 다운로드
def download_google_sheet_as_xlsx(sheet_url, output_filename):
    # 구글 스프레드시트 URL을 XLSX 형식으로 수정
    xlsx_url = sheet_url.replace('/edit?usp=sharing', '/export?format=xlsx')
    
    # 파일 다운로드 요청
    response = requests.get(xlsx_url)
    
    if response.status_code == 200:
        # 다운로드 받은 파일을 지정된 파일명으로 저장
        with open(output_filename, 'wb') as f:
            f.write(response.content)
        print(f"파일 다운로드 완료: {output_filename}")
        print("파일이 있는 위치는 ")
    else:
        print("파일 다운로드 실패:", response.status_code)

# 실행 예시
sheet_url = "https://docs.google.com/spreadsheets/d/1om5aH7l1khkN0wk2ET1E6KzsmXv1_BpzstGBrMfB5Uo/edit?usp=sharing"  # 구글 스프레드시트 링크
output_filename = input("저장할 파일 이름을 입력해주세요: ") + ".xlsx"  # 확장자는 자동으로 붙음





# 프로그램 실행 함수 모음집
download_google_sheet_as_xlsx(sheet_url, output_filename)
check_duplicate_phone_numbers()
